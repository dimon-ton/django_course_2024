from django.shortcuts import render
from django.http import HttpResponse
from .models import *
from openpyxl import Workbook, load_workbook

from datetime import timedelta
from django.utils.timezone import now
from collections import *

from django.utils import timezone

# Create your views here.

def test_pos(req):
    return HttpResponse("<h1>test_pos</h1>")


def AllTable(req):
    try:
        all_table = Table.objects.all()
    except Table.DoesNotExist:
        all_table = None

    context = {"all_table": all_table}

    return render(req, "pos/all-tables.html", context)


def OrdersMenu(req, table_id):
    try:
        table = Table.objects.get(id=table_id)
    except Table.DoesNotExist:
        table = None

 
    all_category = Category.objects.all()
    all_menu = Menu.objects.all()

    new_products = Menu.objects.filter(is_new=True).order_by('-id') # pull the latest ID
    promotions = Menu.objects.filter(is_promotion=True).order_by('-id')

    current_promotions = Promotion.objects.filter(start_date__lte=now(), end_date__gte=now())

    if req.method == "POST":
        data = req.POST.copy()

        order_menu = OrderMenu.objects.create(
            table=table, status="Pending", order_date=timezone.now()
            )
        
        selected_menu = data.getlist("menu")
        counts = data.getlist("count")

        total_price = 0
        items_data = []

        for i, menu_id in enumerate(selected_menu):
            if not menu_id or not str(menu_id).isdigit():
                continue
            try:
                menu = Menu.objects.get(id=menu_id)
            except Menu.DoesNotExist:
                continue


            count = int(counts[i])

            item_total = menu.price * count

            total_price += item_total

            OrderMenuItem.objects.create(
                order_menu=order_menu,
                menu=menu,
                count=count,
                item_total=item_total
            )

            items_data.append({
                "รายการอาหาร": menu.name,
                "จํานวน": count,
                "ราคา": item_total,
                "เวลา": order_menu.order_time
            })

        total_price = float(data.get("total_buyer_price"))
        vat = float(data.get("vat"))
        final_price_with_vat = float(data.get("final_price_with_vat"))

        order_menu.total_buyer_price = total_price
        order_menu.vat = vat
        order_menu.final_price_with_vat = final_price_with_vat

        order_menu.save()

        ########### save to excel ############
        
        file_name = "order.xlsx"

        if os.path.exists(file_name):
            # Load the existing workbook
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            # Create a new workbook if the file doesn't exist
            wb = Workbook()
            ws = wb.active
            # Insert the header if the file is newly created
            header = ["รายการอาหาร", "จำนวน", "ราคา", "เวลา", "โต๊ะ", "vat"]
            ws.append(header)

        # Append the row data
        for row in items_data:
            row_data = list(row.values())
            row_data.extend([table.number, order_menu.vat])
            ws.append(row_data)

        # Save the workbook
        wb.save(file_name)

        context = {"order_menu": order_menu}

        return render(req, "pos/sum-order-menu.html", context)
    
    context = {"table":table, 
               "all_category": all_category, 
               "all_menu": all_menu,
               "new_products": new_products,
               "promotions": promotions,
               "current_promotions": current_promotions
               }
    return render(req, "pos/order-menu.html", context)

def MonthlyOrderSummary(req):
    now = timezone.now()
    start_date = now.replace(day=1)
    end_date = (start_date + timezone.timedelta(days=31)).replace(day=1)

    orders = OrderMenu.objects.filter(order_date__range=[start_date, end_date])

    total_orders = orders.count()
    total_sales = sum(order.final_price_with_vat for order in orders)

    daily_sales = defaultdict(float)

    for order in orders:
        order_date = order.order_date
        daily_sales[order_date] += order.final_price_with_vat



    # graph javascript
    labels = list(map(str, daily_sales.keys()))
    data = list(daily_sales.values())

    context = {"total_orders": total_orders, 
               "total_sales": total_sales, 
               "daily_sales": labels, 
               "daily_sales_data": data,
               }
    
    return render(req, "pos/monthly-order-summary.html", context)